import datetime
import logging
import os

import openpyxl
from openpyxl.styles import Alignment, Font
from django.conf import settings

from django.shortcuts import render
from django.views import View
from django.http import HttpResponse, Http404
from django.utils.decorators import method_decorator
from django.utils.encoding import escape_uri_path
from django.contrib.auth.decorators import login_required

from .models import SubdivisionModel, QuestionModel, QestionResultShowPermission
from .tasks import celery_send_email_to_subdivision_responsible

logging.basicConfig(level=logging.INFO, filename="py_log.log", filemode="w")


# Create your views here.

class IndexView(View):
    """
    Главная страница проекта
    """

    def get(self, request):
        # Список подразделений
        content = {
            'subdivisions': SubdivisionModel.objects.all()
        }
        return render(request, 'anonym_form_app/index.html', content)

    def post(self, request):
        print(f'request.POST: {request.POST}')
        new_question = QuestionModel()
        # Проверка указано ли подразделение
        try:
            new_question.subdivision_id = int(request.POST['subdivision_id'])
        except Exception as e:
            print(e)
        new_question.important_of_question = int(request.POST['important_of_question'])
        new_question.type_of_question = int(request.POST['type_of_question'])
        new_question.question = request.POST['question_text']
        new_question.save()
        celery_send_email_to_subdivision_responsible.delay(new_question.id)
        logging.info(f'Новое обращение: {new_question}')
        content = {
            'new_question': new_question
        }
        return render(request, 'anonym_form_app/ajax/send_question_done.html', content)


class AnalyticView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request):
        user_exist = QestionResultShowPermission.objects.filter(emp__user=request.user).exists()
        if user_exist is False:
            content = {}
            return render(request, 'anonym_form_app/error_permission.html', content)
        tasks = QuestionModel.objects.all().filter(done_flag=False).order_by('-id')
        content = {
            'tasks': tasks,
            'subdivisions': SubdivisionModel.objects.all()
        }
        return render(request, 'anonym_form_app/analytic.html', content)


class AnalyticAllView(View):
    @method_decorator(login_required(login_url='login'))
    def get(self, request):
        user_exist = QestionResultShowPermission.objects.filter(emp__user=request.user).exists()
        if user_exist is False:
            content = {}
            return render(request, 'anonym_form_app/error_permission.html', content)
        tasks = QuestionModel.objects.all().order_by('-id')
        content = {
            'tasks': tasks,
            'subdivisions': SubdivisionModel.objects.all()
        }
        return render(request, 'anonym_form_app/analytic_all.html', content)


class GetTasksWithFilters(View):
    def get(self, request):
        try:
            subdivision_id = int(request.GET['subdivision_id'])
        except:
            subdivision_id = None
        important_of_question = request.GET['important_of_question']
        type_of_question = request.GET['type_of_question']
        status_id = request.GET['status_id']
        tasks = QuestionModel.objects.all().filter(done_flag=False).filter().order_by('-id')
        if subdivision_id == 0:
            tasks = tasks.filter(subdivision_id=None)
        elif subdivision_id:
            tasks = tasks.filter(subdivision_id=subdivision_id)
        if important_of_question:
            tasks = tasks.filter(important_of_question=important_of_question)
        if type_of_question:
            tasks = tasks.filter(type_of_question=type_of_question)
        if status_id:
            tasks = tasks.filter(status=status_id)
        content = {
            'tasks': tasks,
        }
        return render(request, 'anonym_form_app/ajax/tasks_in_work.html', content)


class GetAllTasksWithFilters(View):
    def get(self, request):
        try:
            subdivision_id = int(request.GET['subdivision_id'])
        except:
            subdivision_id = None
        important_of_question = request.GET['important_of_question']
        type_of_question = request.GET['type_of_question']
        status_id = request.GET['status_id']
        tasks = QuestionModel.objects.all().order_by('-id')
        if subdivision_id == 0:
            tasks = tasks.filter(subdivision_id=None)
        elif subdivision_id:
            tasks = tasks.filter(subdivision_id=subdivision_id)
        if important_of_question:
            tasks = tasks.filter(important_of_question=important_of_question)
        if type_of_question:
            tasks = tasks.filter(type_of_question=type_of_question)
        if status_id:
            tasks = tasks.filter(status=status_id)
        content = {
            'tasks': tasks,
        }
        return render(request, 'anonym_form_app/ajax/tasks_all.html', content)


class ModalDetailView(View):
    def get(self, request):
        task_id = request.GET['task_id']
        task = QuestionModel.objects.get(id=task_id)
        len_text = len(task.question)
        try:
            len_description = len(task.description_to_work)
        except Exception as e:
            len_description = 0
        try:
            len_description_cancel = len(task.description_canceled)
        except Exception as e:
            len_description_cancel = 0
        if len_text > 150 or len_description > 150 or len_description_cancel > 150:
            len_flag = True
        else:
            len_flag = False
        content = {
            'task': QuestionModel.objects.get(id=task_id),
            'len_flag': len_flag

        }
        return render(request, 'anonym_form_app/ajax/detail_modal.html', content)

    def post(self, request):
        print(request.POST)
        task_id = int(request.POST['obj_id_for_change_status'])
        task_to_change_status = QuestionModel.objects.get(id=task_id)
        task_to_change_status.status = request.POST['status_id']
        if len(request.POST['WhyCancel']):
            task_to_change_status.description_canceled = request.POST['WhyCancel']
        if len(request.POST['InWorkComment']):
            task_to_change_status.description_to_work = request.POST['InWorkComment']
        if int(request.POST['status_id']) == 2 or int(request.POST['status_id']) == 1:
            task_to_change_status.done_flag = False
            task_to_change_status.start_work_date = datetime.datetime.now()
        elif int(request.POST['status_id']) == 0 or int(request.POST['status_id']) == 3:
            task_to_change_status.done_flag = True
            task_to_change_status.done_date = datetime.datetime.now()
        task_to_change_status.save()
        return HttpResponse(status=200)


class ChangeSubdivisionModalDetailView(View):
    def get(self, request):
        task_id = request.GET['task_id']
        content = {
            'task': QuestionModel.objects.get(id=task_id),
            'subdivisions': SubdivisionModel.objects.all()
        }
        return render(request, 'anonym_form_app/ajax/change_subdivision_modal.html', content)

    def post(self, request):
        task_id = int(request.POST['obj_id_for_change_division'])
        division_id = int(request.POST['division_id'])
        task = QuestionModel.objects.get(id=task_id)
        if division_id == 0:
            task.subdivision_id = None
        else:
            task.subdivision_id = division_id
        task.save()
        return HttpResponse(status=200)


class BlankDetailView(View):
    def get(self, request, pk):
        content = {
            'task': QuestionModel.objects.get(id=pk)
        }
        return render(request, 'anonym_form_app/blank_page.html', content)


class GetXLSReportNewTasks(View):
    def get(self, request):
        file_xlsx_path = os.path.join(settings.BASE_DIR, 'anonym_form_app', 'xlsx', 'report-actual-tasks.xlsx')
        file_path_to_export = os.path.join(settings.BASE_DIR, 'anonym_form_app', 'xlsx', 'report_actual_export.xlsx')
        # wb = openpyxl.Workbook()
        wb = openpyxl.load_workbook(filename=file_xlsx_path, data_only=True)
        page = wb.active
        tasks = QuestionModel.objects.all().filter(done_flag=False).order_by('-id')
        count = 0  # счетчик строк
        for task in tasks:
            row = [
                task.id,
                task.question,
                task.get_status_display(),
                task.get_type_of_question_display(),
                task.get_important_of_question_display(),
                task.date_add,
                task.description_to_work,
                task.start_work_date
            ]
            page.append(row)
            alignment = Alignment(
                horizontal='general',
                vertical='center',
                wrap_text=True,
                shrink_to_fit=True)
            alignment_center_vertical = Alignment(vertical='center', )
            page[f'A{count + 2}'].alignment = alignment
            page[f'B{count + 2}'].alignment = alignment
            page[f'C{count + 2}'].alignment = alignment_center_vertical
            page[f'D{count + 2}'].alignment = alignment_center_vertical
            page[f'E{count + 2}'].alignment = alignment_center_vertical
            page[f'F{count + 2}'].alignment = alignment_center_vertical
            page[f'G{count + 2}'].alignment = alignment
            count += 1
        try:
            wb.save(file_path_to_export)
            print(f'Файл сохранен')
        except Exception as e:
            print(e)
            return False
        if os.path.exists(file_path_to_export):
            # Если файл существует (сохранился)
            with open(file_path_to_export, 'rb') as fh:
                # Установка mimetype для правильной обработки браузером
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                response = HttpResponse(fh.read(), content_type=mime_type)
                # В названии файла устанавливаем номер задачи с пометкой print
                response['Content-Disposition'] = 'attachment; filename=' + escape_uri_path(
                    f'Текущите задачи ({datetime.datetime.now().today()}).xlsx')
                return response
        else:
            return Http404


class GetXLSReportAllTasks(View):
    def get(self, request):
        file_xlsx_path = os.path.join(settings.BASE_DIR, 'anonym_form_app', 'xlsx', 'report-all-tasks.xlsx')
        file_path_to_export = os.path.join(settings.BASE_DIR, 'anonym_form_app', 'xlsx', 'report-all-tasks_export.xlsx')
        # wb = openpyxl.Workbook()
        wb = openpyxl.load_workbook(filename=file_xlsx_path, data_only=True)
        page = wb.active
        tasks = QuestionModel.objects.all().order_by('-id')
        count = 0  # счетчик строк
        for task in tasks:
            row = [
                task.id,
                task.question,
                task.get_status_display(),
                task.get_type_of_question_display(),
                task.get_important_of_question_display(),
                task.date_add,
                task.description_to_work,
                task.start_work_date,
                task.done_date,
                task.description_canceled
            ]
            page.append(row)
            alignment = Alignment(
                horizontal='general',
                vertical='center',
                wrap_text=True,
                shrink_to_fit=True)
            alignment_center_vertical = Alignment(vertical='center', )
            page[f'A{count + 2}'].alignment = alignment
            page[f'B{count + 2}'].alignment = alignment
            page[f'C{count + 2}'].alignment = alignment_center_vertical
            page[f'D{count + 2}'].alignment = alignment_center_vertical
            page[f'E{count + 2}'].alignment = alignment_center_vertical
            page[f'F{count + 2}'].alignment = alignment_center_vertical
            page[f'G{count + 2}'].alignment = alignment
            page[f'H{count + 2}'].alignment = alignment_center_vertical
            page[f'I{count + 2}'].alignment = alignment_center_vertical
            page[f'J{count + 2}'].alignment = alignment
            count += 1
        try:
            wb.save(file_path_to_export)
            print(f'Файл сохранен')
        except Exception as e:
            print(e)
            return False
        if os.path.exists(file_path_to_export):
            # Если файл существует (сохранился)
            with open(file_path_to_export, 'rb') as fh:
                # Установка mimetype для правильной обработки браузером
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                response = HttpResponse(fh.read(), content_type=mime_type)
                # В названии файла устанавливаем номер задачи с пометкой print
                response['Content-Disposition'] = 'attachment; filename=' + escape_uri_path(
                    f'Все задачи от ({datetime.datetime.now().today()}).xlsx')
                return response
        else:
            return Http404
